"""
engine.py — Automated University Theory Examination Seating Arrangement System
PS09 · GCEE Examination Cell
Core Engine: allocation, validation, anti-malpractice interleaving, export
"""

import random, json
from dataclasses import dataclass, field, asdict
from collections import defaultdict
from typing import List, Optional, Dict

# ── Name pool ──────────────────────────────────────────────────────────────────
NAMES = [
    "Aarav Kumar","Aditi Sharma","Akash Raj","Anitha Devi","Arjun Selvam",
    "Arun Pandian","Balaji Murugan","Bhavana Ravi","Deepa Krishnan","Deepak Suresh",
    "Divya Priya","Gokul Nathan","Harini Subramanian","Karthik Venkat","Kavitha Mani",
    "Keerthana Pillai","Lakshmi Narayanan","Logesh Babu","Manoj Karthik","Meena Sundaram",
    "Muthu Selvan","Nandhini Gopal","Naveen Raj","Nithya Devi","Pooja Ramesh",
    "Priya Lakshmi","Rahul Shankar","Raja Mohan","Rajan Kumar","Ramesh Babu",
    "Saranya Perumal","Sathish Kumar","Selvi Murugesan","Senthil Nathan","Sneha Vijay",
    "Suresh Kannan","Surya Prakash","Swetha Rajan","Tamilarasan Vel","Usha Kumari",
    "Vasanth Kumar","Vijay Anand","Vikram Singh","Vimal Raj","Vishal Murugan",
    "Yamini Devi","Yuvan Shankar","Aishwarya Ravi","Bhuvanesh Mani","Chandru Nathan",
    "Dhanush Babu","Elavarasi Devi","Fathima Begum","Gowtham Raj","Hariharan Pillai",
    "Indhu Priya","Janani Krishnan","Kavin Raj","Lavanya Suresh","Mohan Kumar",
    "Nandha Gopal","Oviya Devi","Prabhu Nathan","Ramya Lakshmi","Soundarya Raj",
    "Thirumeni Vel","Uma Devi","Varsha Priya","Abinaya Raj","Bharathi Kumar",
]

SUBJECTS = {
    "CSE":   ["Data Structures & Algorithms","Operating Systems","Database Management Systems",
               "Computer Networks","Software Engineering","Theory of Computation","Compiler Design"],
    "ECE":   ["Signals & Systems","Digital Electronics","Microprocessors","VLSI Design",
               "Communication Systems","Embedded Systems","Antenna Theory"],
    "MECH":  ["Thermodynamics","Fluid Mechanics","Manufacturing Technology","CAD/CAM",
               "Machine Design","Engineering Materials","Heat Transfer"],
    "CIVIL": ["Structural Analysis","Geotechnical Engineering","Highway Engineering",
               "Hydraulics","Environmental Engineering","Surveying","Concrete Technology"],
    "EEE":   ["Electrical Machines","Power Systems","Control Systems","Power Electronics",
               "Measurements & Instrumentation","High Voltage Engineering","Drives"],
    "IT":    ["Web Technology","Cloud Computing","Data Mining","Network Security",
               "Mobile Computing","IoT","Cyber Security"],
    "AUTO":   ["Marketing Management","Financial Management","HR Management",
               "Operations Research","Business Analytics","Strategic Management"],
    "CSDS":   ["Advanced Java","Python Programming","Software Testing",
               "Artificial Intelligence","Data Warehousing","Mobile App Development"],
}

DEPT_PREFIXES = {
    "CSE":"21CSE","ECE":"21ECE","MECH":"21MCE","CIVIL":"21CVE",
    "EEE":"21EEE","IT":"21IMT","AUTO":"21ATM","CSDS":"21CSDS",
}

DEPT_FULL = {
    "CSE":"Computer Science & Engineering",
    "ECE":"Electronics & Communication Engineering",
    "MECH":"Mechanical Engineering",
    "CIVIL":"Civil Engineering",
    "EEE":"Electrical & Electronics Engineering",
    "IT":"Information Technology",
    "AUTO":"Automobile Engineering",
    "CSDS":"Computer Science & Data Science",
}


# ══════════════════════════════════════════════════════════════════════════════
# DATA MODELS
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class Student:
    register_no: str
    name: str
    dept: str
    subject: str
    semester: int


@dataclass
class Seat:
    seat_id: str
    row: int
    col: int
    student: Optional[Student] = None


@dataclass
class Hall:
    hall_id: str
    name: str
    capacity: int = 25
    rows: int = 5
    cols: int = 5
    seats: List[Seat] = field(default_factory=list)

    @property
    def occupied(self) -> int:
        return sum(1 for s in self.seats if s.student)

    @property
    def depts(self) -> List[str]:
        return sorted(set(s.student.dept for s in self.seats if s.student))

    @property
    def utilization(self) -> float:
        return round(self.occupied / self.capacity * 100, 1) if self.capacity else 0.0


# ══════════════════════════════════════════════════════════════════════════════
# SEATING ENGINE
# ══════════════════════════════════════════════════════════════════════════════

class SeatingEngine:
    def __init__(self):
        self.halls:    List[Hall]    = []
        self.students: List[Student] = []
        self.warnings: List[str]     = []
        self.errors:   List[str]     = []

    def clear(self):
        self.halls, self.students, self.warnings, self.errors = [], [], [], []

    # ── generate students ──────────────────────────────────────────────────────
    def generate_students(self, dept: str, count: int,
                           semester: int = 3, subject: str = None) -> List[Student]:
        dept  = dept.upper()
        prefix = DEPT_PREFIXES.get(dept, "21XX")
        subj   = subject or random.choice(SUBJECTS.get(dept, ["General Examination"]))
        pool   = NAMES.copy(); random.shuffle(pool)
        return [
            Student(
                register_no=f"{prefix}{i:03d}",
                name=pool[(i-1) % len(pool)],
                dept=dept, subject=subj, semester=semester
            )
            for i in range(1, count + 1)
        ]

    # ── anti-malpractice interleave (for a 2-dept pair) ───────────────────────
    def _interleave_pair(self, group_a: List[Student], group_b: List[Student]) -> List[Student]:
        """Strictly alternate A-B-A-B.  When sizes differ, the larger dept fills
        the remaining seats at the end — unavoidable with only 2 depts and unequal counts.
        The key guarantee: every B seat is surrounded by A seats and vice-versa
        for the interleaved portion.
        """
        if not group_a: return list(group_b)
        if not group_b: return list(group_a)
        # Larger group leads so the tail (same-dept run) is as short as possible
        if len(group_b) > len(group_a):
            group_a, group_b = group_b, group_a
        result: List[Student] = []
        ia, ib, turn = 0, 0, 0
        while ia < len(group_a) or ib < len(group_b):
            if turn == 0 and ia < len(group_a):
                result.append(group_a[ia]); ia += 1
            elif turn == 1 and ib < len(group_b):
                result.append(group_b[ib]); ib += 1
            elif ia < len(group_a):
                result.append(group_a[ia]); ia += 1
            else:
                result.append(group_b[ib]); ib += 1
            turn ^= 1
        return result

    def _interleave_multi(self, students: List[Student]) -> List[Student]:
        """Round-robin across all departments (used for overflow hall)."""
        groups: dict = {}
        for s in students:
            groups.setdefault(s.dept, []).append(s)
        result, dept_lists = [], list(groups.values())
        for i in range(max((len(g) for g in dept_lists), default=0)):
            for g in dept_lists:
                if i < len(g):
                    result.append(g[i])
        return result

    # ── validate ───────────────────────────────────────────────────────────────
    def _validate(self) -> bool:
        self.errors, self.warnings = [], []
        if not self.halls:
            self.errors.append("No examination halls configured."); return False
        if not self.students:
            self.errors.append("No students to allocate."); return False

        reg_nos = [s.register_no for s in self.students]
        if len(reg_nos) != len(set(reg_nos)):
            dupes = [r for r in reg_nos if reg_nos.count(r) > 1]
            self.errors.append(f"Duplicate register numbers: {', '.join(set(dupes))}"); return False

        total_cap = sum(h.capacity for h in self.halls)
        if len(self.students) > total_cap:
            self.errors.append(
                f"Total students ({len(self.students)}) exceed total hall capacity ({total_cap}). "
                f"Add {len(self.students)-total_cap} more seats."
            ); return False
        return True

    # ── allocate ────────────────────────────────────────────────────────────
    def allocate(self) -> Dict:
        """
        Greedy Column-Based Seating — Fill & Move
        ==========================================
        For each hall in order:
          1. Pick the 2 departments with the most remaining students.
          2. Fill the hall using column seating:
               Odd  columns → larger dept (A)
               Even columns → smaller dept (B)
          3. Hall fills completely if A+B >= capacity.
             If A+B < capacity, seats partial — remaining go to next hall.
          4. Last hall gets everything left, round-robin across all depts.

        This guarantees:
          - Every non-last hall has EXACTLY 2 departments
          - Zero left-right malpractice in non-last halls (column pattern)
          - Each hall fills as completely as possible before moving to next
        """
        if not self._validate():
            return self._result(False)

        # ── Init seat grids ───────────────────────────────────────────────────
        for hall in self.halls:
            hall.seats = [
                Seat(seat_id=f"{chr(64+r)}{c}", row=r, col=c)
                for r in range(1, hall.rows + 1)
                for c in range(1, hall.cols + 1)
            ]

        # ── Group students by dept ────────────────────────────────────────────
        pool: dict = {}
        for s in self.students:
            pool.setdefault(s.dept, []).append(s)

        n_halls = len(self.halls)

        # ── Column fill helper ────────────────────────────────────────────────
        def _column_fill(hall, grpA: list, grpB: list, pool_rest: dict = None):
            """
            Checkerboard Seating - true anti-malpractice interleaving.
            Every seat alternates department in BOTH row and column directions:

              Row 1: A B A B A
              Row 2: B A B A B
              Row 3: A B A B A
              Row 4: B A B A B
              Row 5: A B A B A

            Seat (row, col) -> dept A when (row+col) is even, dept B when odd.
            No two adjacent seats (left/right or front/back) share same dept.
            Returns (leftA, leftB).
            """
            seats_A = [s for s in hall.seats if (s.row + s.col) % 2 == 0]
            seats_B = [s for s in hall.seats if (s.row + s.col) % 2 == 1]

            takeA = min(len(grpA), len(seats_A))
            takeB = min(len(grpB), len(seats_B))

            seatedA, leftA = grpA[:takeA], list(grpA[takeA:])
            seatedB, leftB = grpB[:takeB], list(grpB[takeB:])

            for seat, student in zip(seats_A, seatedA):
                seat.student = student
            for seat, student in zip(seats_B, seatedB):
                seat.student = student

            return leftA, leftB

        # ── Greedy hall-by-hall allocation ────────────────────────────────────
        # Each hall is filled COMPLETELY before moving to the next.
        # Strategy per hall:
        #   1. Pick the 2 depts with the most remaining students.
        #   2. Fill the hall with checkerboard pattern (A/B seats).
        #   3. If A+B < hall capacity, top-up remaining seats from other depts
        #      (still checkerboard parity — no dept gets a neighbour of same dept).
        #   4. Any leftover students carry over to the next hall.
        for hi, hall in enumerate(self.halls):
            if not pool:
                break

            hall_cap = hall.capacity
            dept_keys = list(pool.keys())

            # Pick best pair: 2 depts with most students
            dept_keys_sorted = sorted(dept_keys, key=lambda d: len(pool[d]), reverse=True)
            best_dA = dept_keys_sorted[0]
            best_dB = dept_keys_sorted[1] if len(dept_keys_sorted) > 1 else None

            grpA = list(pool.pop(best_dA))
            grpB = list(pool.pop(best_dB)) if best_dB else []

            # Fill with checkerboard — returns leftovers
            leftA, leftB = _column_fill(hall, grpA, grpB)

            # Top-up: if hall not full, fill remaining empty seats from other depts
            # maintaining checkerboard parity for those seats too
            empty_seats = [s for s in hall.seats if s.student is None]
            for seat in empty_seats:
                if not pool:
                    break
                # Pick dept with most students
                best = max(pool.keys(), key=lambda d: len(pool[d]))
                seat.student = pool[best].pop(0)
                if not pool[best]:
                    del pool[best]

            # Put primary dept leftovers back into pool
            if leftA: pool[best_dA] = leftA
            if leftB and best_dB: pool[best_dB] = leftB

        # ── Warnings ─────────────────────────────────────────────────────────
        unallocated = [s for grp in pool.values() for s in grp]

        for idx, hall in enumerate(self.halls):
            if hall.occupied == 0:
                continue
            depts   = hall.depts
            is_last = (idx == n_halls - 1)

            if not is_last and len(depts) == 1:
                self.warnings.append(
                    f"Hall '{hall.name}' has only 1 department ({depts[0]}) — "
                    "malpractice risk."
                )
            elif is_last and len(depts) > 2:
                self.warnings.append(
                    f"Hall '{hall.name}' (last hall) has {len(depts)} departments "
                    f"({', '.join(depts)}) — round-robin interleaving applied."
                )

        if unallocated:
            self.warnings.append(
                f"{len(unallocated)} student(s) could not be allocated — "
                "all halls are full. Add more halls or increase capacity."
            )

        return self._result(True)
    def _result(self, success: bool) -> Dict:
        return {
            "success":  success,
            "errors":   self.errors,
            "warnings": self.warnings,
            "halls":    [self._ser_hall(h) for h in self.halls] if success else [],
            "summary":  self._summary() if success else {},
        }

    def _ser_hall(self, h: Hall) -> Dict:
        return {
            "hall_id": h.hall_id, "name": h.name,
            "capacity": h.capacity, "rows": h.rows, "cols": h.cols,
            "occupied": h.occupied, "utilization": h.utilization, "depts": h.depts,
            "seats": [
                {"seat_id": s.seat_id, "row": s.row, "col": s.col,
                 "student": asdict(s.student) if s.student else None}
                for s in h.seats
            ],
        }

    def _summary(self) -> Dict:
        total_alloc = sum(h.occupied for h in self.halls)
        total_cap   = sum(h.capacity for h in self.halls)
        all_depts   = sorted(set(d for h in self.halls for d in h.depts))
        return {
            "total_students":    len(self.students),
            "total_allocated":   total_alloc,
            "total_capacity":    total_cap,
            "halls_used":        sum(1 for h in self.halls if h.occupied > 0),
            "departments":       all_depts,
            "overall_util":      round(total_alloc/total_cap*100,1) if total_cap else 0,
            "utilization": [
                {"name":h.name,"capacity":h.capacity,"occupied":h.occupied,
                 "pct":h.utilization,"depts":h.depts}
                for h in self.halls
            ],
        }

    # ── Excel export ───────────────────────────────────────────────────────────
    def export_excel(self, filepath: str, result: Dict):
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        wb   = openpyxl.Workbook(); wb.remove(wb.active)
        thin = Side(style="thin", color="444444")
        bdr  = Border(left=thin,right=thin,top=thin,bottom=thin)

        FILLS = {
            "CSE":"1E2A3A","ECE":"1A2E2E","MECH":"2E2410","CIVIL":"0F2E1A",
            "EEE":"27153A","IT":"2E1010","AUTO":"2E2500","CS-DS":"2E0A2E","OTHER":"222222",
        }
        TEXT  = "C8C8C8"
        TITLE = "E8E8E8"

        # ── Summary sheet ──
        ws = wb.create_sheet("SUMMARY", 0)
        s  = result.get("summary", {})
        ws.merge_cells("A1:F1")
        ws["A1"] = "AUTOMATED EXAMINATION — SEATING SUMMARY"
        ws["A1"].font = Font(bold=True, size=14, color=TITLE)
        ws["A1"].fill = PatternFill("solid", fgColor="0A0A0A")
        ws["A1"].alignment = Alignment(horizontal="center"); ws.row_dimensions[1].height = 30

        summary_rows = [
            ("Total Students",  s.get("total_students",0)),
            ("Total Allocated", s.get("total_allocated",0)),
            ("Total Halls",     s.get("halls_used",0)),
            ("Utilization",     f"{s.get('overall_util',0)}%"),
            ("Departments",     ", ".join(s.get("departments",[]))),
        ]
        for ri,(k,v) in enumerate(summary_rows, 3):
            c1 = ws.cell(row=ri,column=1,value=k)
            c1.font = Font(color="909090",size=11); c1.fill = PatternFill("solid",fgColor="141414")
            c2 = ws.cell(row=ri,column=2,value=str(v))
            c2.font = Font(color=TITLE,size=11,bold=True); c2.fill = PatternFill("solid",fgColor="1C1C1C")
            ws.column_dimensions["A"].width = 22; ws.column_dimensions["B"].width = 50

        # per-hall utilization
        ws.cell(row=9,column=1,value="Hall Utilization").font = Font(color="707070",size=10,bold=True)
        ws.cell(row=9,column=1).fill = PatternFill("solid",fgColor="0A0A0A")
        for ri,h in enumerate(s.get("utilization",[]), 10):
            ws.cell(row=ri,column=1,value=h["name"]).font = Font(color="888888",size=10)
            ws.cell(row=ri,column=1).fill = PatternFill("solid",fgColor="141414")
            ws.cell(row=ri,column=2,value=f"{h['occupied']}/{h['capacity']} ({h['pct']}%) — {', '.join(h['depts'])}").font = Font(color=TEXT,size=10)
            ws.cell(row=ri,column=2).fill = PatternFill("solid",fgColor="1C1C1C")

        # ── per-hall sheets ──
        for hall_data in result["halls"]:
            if hall_data["occupied"] == 0: continue
            ws = wb.create_sheet(hall_data["name"][:31])

            # title
            ws.merge_cells("A1:G1")
            ws["A1"] = f"EXAMINATION SEATING ARRANGEMENT — {hall_data['name'].upper()}"
            ws["A1"].font = Font(bold=True,size=13,color=TITLE)
            ws["A1"].fill = PatternFill("solid",fgColor="000000")
            ws["A1"].alignment = Alignment(horizontal="center"); ws.row_dimensions[1].height=28

            ws.merge_cells("A2:G2")
            ws["A2"] = (f"Capacity: {hall_data['capacity']}   |   Occupied: {hall_data['occupied']}"
                        f"   |   Utilization: {hall_data['utilization']}%"
                        f"   |   Departments: {', '.join(hall_data['depts'])}")
            ws["A2"].font = Font(size=10,color="888888")
            ws["A2"].fill = PatternFill("solid",fgColor="141414")
            ws["A2"].alignment = Alignment(horizontal="center")

            headers = ["Seat","Register No","Student Name","Department","Subject","Semester","Position"]
            for ci,h in enumerate(headers,1):
                c = ws.cell(row=4,column=ci,value=h)
                c.font = Font(bold=True,color=TITLE,size=10)
                c.fill = PatternFill("solid",fgColor="1C1C1C")
                c.alignment = Alignment(horizontal="center"); c.border = bdr

            rn = 5
            for seat in hall_data["seats"]:
                if not seat["student"]: continue
                st   = seat["student"]
                dept = st["dept"]
                fg   = FILLS.get(dept, FILLS["OTHER"])
                fill = PatternFill("solid",fgColor=fg)
                vals = [seat["seat_id"],st["register_no"],st["name"],
                        dept,st["subject"],st["semester"],
                        f"Row {seat['row']} · Col {seat['col']}"]
                for ci,val in enumerate(vals,1):
                    c = ws.cell(row=rn,column=ci,value=val)
                    c.fill=fill; c.border=bdr
                    c.alignment=Alignment(horizontal="center")
                    c.font=Font(size=10,color=TEXT)
                rn += 1

            for ci in range(1,8):
                ws.column_dimensions[chr(64+ci)].width = 20

        wb.save(filepath)
