"""
app.py — Automated Examination Seating Arrangement System
GCEE Examination Cell
Run: python app.py   →   http://localhost:5000
"""

import os, json, uuid, re, tempfile, traceback
import urllib.request as _urllib
import urllib.error
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file, session
from engine import SeatingEngine, Hall, SUBJECTS, DEPT_PREFIXES, DEPT_FULL, NAMES

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "gcee_examcell_ps09_2024_secret")

_engines: dict = {}   # session_id → SeatingEngine

def _engine(sid: str) -> SeatingEngine:
    if sid not in _engines:
        _engines[sid] = SeatingEngine()
    return _engines[sid]

def _sid() -> str:
    if "sid" not in session:
        session["sid"] = str(uuid.uuid4())
    return session["sid"]


# ── global JSON error handlers ─────────────────────────────────────────────────
@app.errorhandler(404)
def e404(e): return jsonify({"error": "Not found"}), 404
@app.errorhandler(500)
def e500(e): return jsonify({"error": "Server error", "detail": str(e)}), 500
@app.errorhandler(Exception)
def eAny(e): traceback.print_exc(); return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════════════════════
# AI LAYER  (uses stdlib urllib — no naming clash with Flask's `request`)
# ══════════════════════════════════════════════════════════════════════════════
SYSTEM_PROMPT = """You are the AI for an Automated University Theory Examination Seating Arrangement System
(PS09, GCEE Examination Cell).

Constraints the system enforces:
- Max 25 students per hall (default capacity)
- EXACTLY 2 departments per hall (for easy question paper distribution)
- Anti-malpractice: students interleaved so adjacent seats differ in department (A-B-A-B pattern)
- Students from different semesters CAN sit together — no semester restriction
- Last hall may contain overflow from 2+ departments if needed
- Seat IDs: A1–E5 for a 5×5 hall grid

Your job: extract structured JSON from the user's natural language description.

Always respond with:
1. One friendly sentence confirming what you understood
2. A JSON block wrapped EXACTLY as: <<<JSON { ... } >>>

JSON format:
{
  "action": "generate",
  "exam_name": "Mid Semester Examination",
  "halls": [{"name":"Hall A","capacity":25,"rows":5,"cols":5}],
  "departments": [{"dept":"CSE","count":20,"semester":3,"subject":"Data Structures"}]
}

Rules:
- Dept codes: CSE, ECE, MECH, CIVIL, EEE, IT, AUTO, CS-DS
- Default capacity=25, rows=5, cols=5, semester=3
- Distribute evenly if only total is given
- action="info" for stats/questions (omit halls & departments)"""


def _call_ai(messages: list) -> str:
    key = os.environ.get("ANTHROPIC_API_KEY","").strip()
    if not key:
        return _fallback(messages[-1]["content"])
    try:
        body = json.dumps({
            "model":"claude-sonnet-4-20250514","max_tokens":1500,
            "system":SYSTEM_PROMPT,"messages":messages[-20:],
        }).encode()
        req = _urllib.Request(
            "https://api.anthropic.com/v1/messages", data=body, method="POST",
            headers={"x-api-key":key,"anthropic-version":"2023-06-01",
                     "content-type":"application/json"})
        with _urllib.urlopen(req, timeout=30) as r:
            return json.loads(r.read())["content"][0]["text"]
    except Exception as ex:
        print(f"[AI error] {ex}")
        return _fallback(messages[-1]["content"])


def _fallback(msg: str) -> str:
    """Robust regex fallback — works without any API key."""
    m = msg.lower()
    hm = re.search(r'(\d+)\s*hall', m)
    n_halls = int(hm.group(1)) if hm else 3
    def _hall_name(i):
        L = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        return "Hall " + (L[i] if i < 26 else L[i//26-1]+L[i%26])
    halls = [{"name":_hall_name(i),"capacity":25,"rows":5,"cols":5} for i in range(n_halls)]

    KW = {"cse":"CSE","computer science":"CSE","ece":"ECE","electronics":"ECE",
          "mech":"MECH","mechanical":"MECH","civil":"CIVIL","eee":"EEE",
          "electrical":"EEE"," it ":"IT","information technology":"IT",
          "auto":"AUTO","automobile":"AUTO","csds":"Computer Science & Data Science","CSDS":"Computer Science & Data Science"}
    depts, seen = [], set()
    for kw, code in KW.items():
        if kw.strip() in m and code not in seen:
            pat = rf'(\d+)\s*{re.escape(kw.strip())}|{re.escape(kw.strip())}\s*[:\-]?\s*(\d+)'
            dm  = re.search(pat, m)
            cnt = int(dm.group(1) or dm.group(2)) if dm else 0
            if cnt == 0:
                tm = re.search(r'(\d+)\s*student', m)
                cnt = max(10, int(tm.group(1))//5) if tm else 12
            depts.append({"dept":code,"count":cnt,"semester":3}); seen.add(code)

    if not depts:
        tm = re.search(r'(\d+)\s*student', m)
        base = max(10, int(tm.group(1))//5) if tm else 12
        for d in ["CSE","ECE","MECH","CIVIL","EEE","AUTO"]:
            depts.append({"dept":d,"count":base,"semester":3})

    total = sum(d["count"] for d in depts)
    # Auto-calculate halls if not explicitly mentioned: ceil(depts/2)
    if not hm:
        import math
        n_halls = max(3, math.ceil(len(depts) / 2))
        halls = [{"name":_hall_name(i),"capacity":25,"rows":5,"cols":5} for i in range(n_halls)]
    payload = {"action":"generate","exam_name":"Mid Semester Examination",
               "halls":halls,"departments":depts}
    return (f"Arranging {total} students from {len(depts)} department(s) across {n_halls} hall(s).\n"
            f"<<<JSON\n{json.dumps(payload,indent=2)}\n>>>")


def _parse_ai(text: str):
    match = re.search(r'<<<JSON\s*([\s\S]*?)\s*>>>', text)
    natural = re.sub(r'<<<JSON[\s\S]*?>>>', '', text).strip()
    data = None
    if match:
        try: data = json.loads(match.group(1))
        except Exception as ex: print(f"[JSON parse] {ex}")
    return natural, data


# ══════════════════════════════════════════════════════════════════════════════
# ROUTES
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/")
def index():
    _sid()
    return render_template("index.html",
        has_api_key=bool(os.environ.get("ANTHROPIC_API_KEY","").strip()),
        dept_list=list(DEPT_PREFIXES.keys()),
        dept_full=DEPT_FULL,
        subjects=SUBJECTS,
    )

# ── AI chat ────────────────────────────────────────────────────────────────────
@app.route("/api/chat", methods=["POST"])
def chat():
    try:
        sid  = _sid()
        body = request.get_json(silent=True) or {}
        msg  = (body.get("message") or "").strip()
        hist = body.get("history") or []
        if not msg: return jsonify({"error":"Empty message"}), 400

        hist.append({"role":"user","content":msg})
        ai_raw = _call_ai(hist)
        natural, data = _parse_ai(ai_raw)

        result = None
        if data and data.get("action") in ("generate","modify"):
            eng = _engine(sid); eng.clear()
            for h in data.get("halls",[]):
                eng.halls.append(Hall(
                    hall_id=f"H{len(eng.halls)+1}",
                    name=str(h.get("name",f"Hall {len(eng.halls)+1}")),
                    capacity=int(h.get("capacity",25)),
                    rows=int(h.get("rows",5)), cols=int(h.get("cols",5)),
                ))
            for di in data.get("departments",[]):
                eng.students.extend(eng.generate_students(
                    dept=str(di.get("dept","CSE")).upper(),
                    count=int(di.get("count",10)),
                    semester=int(di.get("semester",3)),
                    subject=di.get("subject"),
                ))
            result = eng.allocate(); _engines[sid] = eng

        return jsonify({"reply":natural or "Arrangement generated.","data":data,"result":result})
    except Exception as ex:
        traceback.print_exc()
        return jsonify({"error":str(ex),"reply":f"Error: {ex}"}), 500


# ── Manual entry (the key new feature) ────────────────────────────────────────
@app.route("/api/manual", methods=["POST"])
def manual():
    """
    Accept structured form data:
    {
      "halls": [{"name":"Hall A","capacity":25,"rows":5,"cols":5}],
      "students": [{"register_no":"21CS001","name":"Arjun","dept":"CSE","subject":"DS","semester":3}]
    }
    OR departments shorthand:
    {
      "halls": [...],
      "departments": [{"dept":"CSE","count":20,"semester":3,"subject":"DS"}]
    }
    """
    try:
        sid  = _sid()
        body = request.get_json(silent=True) or {}
        halls_data = body.get("halls", [])
        students_data = body.get("students", [])
        depts_data = body.get("departments", [])

        if not halls_data:
            return jsonify({"success":False,"errors":["No halls provided."]}), 400
        if not students_data and not depts_data:
            return jsonify({"success":False,"errors":["No student or department data provided."]}), 400

        eng = _engine(sid); eng.clear()

        # build halls — capacity is always rows×cols (grid-derived)
        for i, h in enumerate(halls_data):
            rows = int(h.get("rows", 5))
            cols = int(h.get("cols", 5))
            cap  = rows * cols   # always authoritative
            eng.halls.append(Hall(
                hall_id=f"H{i+1}",
                name=str(h.get("name", ("Hall "+("ABCDEFGHIJKLMNOPQRSTUVWXYZ"[i] if i<26 else "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[i//26-1]+"ABCDEFGHIJKLMNOPQRSTUVWXYZ"[i%26])))),
                capacity=cap, rows=rows, cols=cols,
            ))

        # manual students (individual rows)
        if students_data:
            seen_regs = set()
            errors = []
            for i, s in enumerate(students_data, 1):
                reg = str(s.get("register_no","")).strip()
                if not reg:
                    errors.append(f"Row {i}: Missing register number."); continue
                if reg in seen_regs:
                    errors.append(f"Row {i}: Duplicate register number '{reg}'."); continue
                seen_regs.add(reg)
                eng.students.append(
                    __import__('engine').Student(
                        register_no=reg,
                        name=str(s.get("name","Unknown")).strip(),
                        dept=str(s.get("dept","CSE")).upper().strip(),
                        subject=str(s.get("subject","General")).strip(),
                        semester=int(s.get("semester",3)),
                    )
                )
            if errors:
                return jsonify({"success":False,"errors":errors}), 400

        # department bulk generation
        elif depts_data:
            for di in depts_data:
                eng.students.extend(eng.generate_students(
                    dept=str(di.get("dept","CSE")).upper(),
                    count=int(di.get("count",10)),
                    semester=int(di.get("semester",3)),
                    subject=di.get("subject"),
                ))

        result = eng.allocate()
        _engines[sid] = eng
        return jsonify(result)

    except Exception as ex:
        traceback.print_exc()
        return jsonify({"success":False,"errors":[str(ex)]}), 500


# ── reshuffle ──────────────────────────────────────────────────────────────────
@app.route("/api/reshuffle")
def reshuffle():
    """
    Reshuffle seats WITHIN each hall only.
    - Total students seated stays exactly the same.
    - Each hall keeps the same students it already has.
    - Only the seat positions of students within each hall are shuffled.
    - Anti-malpractice interleaving is re-applied after shuffle.
    """
    try:
        sid = _sid()
        if sid not in _engines:
            return jsonify({"error":"No arrangement found. Generate one first."}), 404

        import random

        eng = _engines[sid]

        # Verify halls are already set up (allocate was called before)
        if not eng.halls or not any(h.seats for h in eng.halls):
            return jsonify({"error": "No arrangement found. Generate one first."}), 404

        # For each hall: collect currently-seated students,
        # shuffle them, re-interleave by dept, put back in the SAME seats.
        # Total seated count never changes — same students, different seat positions.
        for hall_obj in eng.halls:
            # Collect all students currently seated in this hall
            seated = [s.student for s in hall_obj.seats if s.student is not None]
            if not seated:
                continue

            # Shuffle within each dept group first (so individual order varies)
            groups: dict = {}
            for s in seated:
                groups.setdefault(s.dept, []).append(s)
            for grp in groups.values():
                random.shuffle(grp)

            # Re-interleave by dept to restore anti-malpractice alternation
            if len(groups) == 1:
                reordered = list(groups.values())[0]
            elif len(groups) == 2:
                depts = list(groups.keys())
                reordered = eng._interleave_pair(groups[depts[0]], groups[depts[1]])
            else:
                reordered = eng._interleave_multi(seated)

            # Re-assign to the same occupied seat positions — empty seats stay empty
            occupied_seats = [s for s in hall_obj.seats if s.student is not None]
            for seat_obj, student in zip(occupied_seats, reordered):
                seat_obj.student = student

        # Serialize current hall state WITHOUT re-running allocate()
        # Re-serialize the current hall state directly
        from dataclasses import asdict
        halls_out = []
        for h in eng.halls:
            halls_out.append({
                "hall_id":    h.hall_id,
                "name":       h.name,
                "capacity":   h.capacity,
                "rows":       h.rows,
                "cols":       h.cols,
                "occupied":   h.occupied,
                "utilization": h.utilization,
                "depts":      h.depts,
                "seats": [
                    {"seat_id": s.seat_id, "row": s.row, "col": s.col,
                     "student": asdict(s.student) if s.student else None}
                    for s in h.seats
                ]
            })

        total_alloc = sum(h.occupied for h in eng.halls)
        total_cap   = sum(h.capacity for h in eng.halls)
        all_depts   = sorted(set(d for h in eng.halls for d in h.depts))

        result_out = {
            "success":  True,
            "errors":   [],
            "warnings": [],
            "halls":    halls_out,
            "summary": {
                "total_students":  len(eng.students),
                "total_allocated": total_alloc,
                "total_capacity":  total_cap,
                "halls_used":      sum(1 for h in eng.halls if h.occupied > 0),
                "departments":     all_depts,
                "overall_util":    round(total_alloc / total_cap * 100, 1) if total_cap else 0,
                "utilization": [
                    {"name": h.name, "capacity": h.capacity, "occupied": h.occupied,
                     "pct": h.utilization, "depts": h.depts}
                    for h in eng.halls
                ]
            }
        }

        return jsonify({"result": result_out,
                        "reply": f"Seats reshuffled within each hall. {total_alloc} students remain seated."})
    except Exception as ex:
        traceback.print_exc(); return jsonify({"error": str(ex)}), 500


# ── export ─────────────────────────────────────────────────────────────────────
@app.route("/api/export/<fmt>")
def export(fmt):
    try:
        sid = _sid()
        if sid not in _engines:
            return jsonify({"error":"No arrangement found."}), 404
        eng = _engines[sid]; result = eng.allocate()
        ts  = datetime.now().strftime("%Y%m%d_%H%M%S")

        if fmt == "excel":
            tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            eng.export_excel(tmp.name, result)
            return send_file(tmp.name, as_attachment=True,
                download_name=f"seating_arrangement_{ts}.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        elif fmt == "csv":
            import io, csv as _csv
            buf = io.StringIO()
            w   = _csv.writer(buf)
            w.writerow(["Hall","Seat","Register No","Name","Department","Subject","Semester"])
            for h in result["halls"]:
                for s in h["seats"]:
                    if s["student"]:
                        st = s["student"]
                        w.writerow([h["name"],s["seat_id"],st["register_no"],
                                    st["name"],st["dept"],st["subject"],st["semester"]])
            buf.seek(0)
            return send_file(io.BytesIO(buf.getvalue().encode()), as_attachment=True,
                download_name=f"seating_arrangement_{ts}.csv", mimetype="text/csv")

        return jsonify({"error":f"Unknown format: {fmt}"}), 400
    except Exception as ex:
        traceback.print_exc(); return jsonify({"error":str(ex)}), 500


# ── misc ───────────────────────────────────────────────────────────────────────
@app.route("/api/clear")
def clear():
    sid = _sid()
    if sid in _engines: del _engines[sid]
    return jsonify({"ok":True})

@app.route("/api/status")
def status():
    return jsonify({
        "status":"running",
        "api_key":bool(os.environ.get("ANTHROPIC_API_KEY","").strip()),
        "sessions":len(_engines),
    })

@app.route("/api/subjects/<dept>")
def get_subjects(dept):
    return jsonify(SUBJECTS.get(dept.upper(), []))


# ── bulk upload: parse uploaded CSV/Excel file ─────────────────────────────────
@app.route("/api/upload/students", methods=["POST"])
def upload_students():
    """
    Parse an uploaded CSV or Excel file of individual students.
    Expected columns (any order, case-insensitive):
      register_no, name, dept/department, subject, semester/sem
    Returns: {"rows": [...], "errors": [...], "count": N}
    """
    try:
        if "file" not in request.files:
            return jsonify({"error": "No file uploaded"}), 400
        f    = request.files["file"]
        name = f.filename.lower()
        rows, errors = [], []

        if name.endswith(".csv"):
            import io, csv as _csv
            text    = f.read().decode("utf-8-sig", errors="replace")
            reader  = _csv.DictReader(io.StringIO(text))
            raw_rows = list(reader)

        elif name.endswith((".xlsx", ".xls")):
            import openpyxl, io
            wb   = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
            ws   = wb.active
            hdrs = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            raw_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                raw_rows.append({hdrs[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
        else:
            return jsonify({"error": "Only .csv, .xlsx or .xls files are supported"}), 400

        # normalise column names
        COL_MAP = {
            "register_no":"register_no","reg_no":"register_no","regno":"register_no",
            "register":"register_no","roll_no":"register_no","rollno":"register_no",
            "name":"name","student_name":"name","studentname":"name",
            "dept":"dept","department":"dept","branch":"dept",
            "subject":"subject","sub":"subject","paper":"subject",
            "semester":"semester","sem":"semester","year":"semester",
        }

        seen_regs = set()
        for i, raw in enumerate(raw_rows, 1):
            norm = {COL_MAP.get(k.strip().lower(), k.strip().lower()): v
                    for k, v in raw.items() if k.strip()}
            reg  = norm.get("register_no","").strip()
            name_val = norm.get("name","").strip() or f"Student {i}"
            dept_val = norm.get("dept","CSE").strip().upper()
            subj_val = norm.get("subject","").strip()
            sem_val  = norm.get("semester","3")

            if not reg:
                errors.append(f"Row {i}: Missing register number — skipped."); continue
            if reg in seen_regs:
                errors.append(f"Row {i}: Duplicate register number '{reg}' — skipped."); continue
            seen_regs.add(reg)

            try: sem_int = int(float(sem_val))
            except: sem_int = 3

            if dept_val not in DEPT_PREFIXES:
                errors.append(f"Row {i}: Unknown department '{dept_val}' — defaulted to CSE.")
                dept_val = "CSE"

            if not subj_val:
                import random
                subj_val = random.choice(SUBJECTS.get(dept_val, ["General Examination"]))

            rows.append({"register_no":reg,"name":name_val,
                         "dept":dept_val,"subject":subj_val,"semester":sem_int})

        return jsonify({"rows": rows, "errors": errors, "count": len(rows)})

    except Exception as ex:
        traceback.print_exc()
        return jsonify({"error": str(ex)}), 500


@app.route("/api/upload/departments", methods=["POST"])
def upload_departments():
    """
    Parse an uploaded CSV or Excel file of department counts.
    Expected columns: dept/department, count/students, semester/sem, subject (optional)
    Returns: {"rows": [...], "errors": [...]}
    """
    try:
        if "file" not in request.files:
            return jsonify({"error": "No file uploaded"}), 400
        f    = request.files["file"]
        name = f.filename.lower()

        if name.endswith(".csv"):
            import io, csv as _csv
            text   = f.read().decode("utf-8-sig", errors="replace")
            reader = _csv.DictReader(io.StringIO(text))
            raw_rows = list(reader)
        elif name.endswith((".xlsx", ".xls")):
            import openpyxl, io
            wb   = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
            ws   = wb.active
            hdrs = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            raw_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                raw_rows.append({hdrs[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
        else:
            return jsonify({"error": "Only .csv or .xlsx files are supported"}), 400

        COL_MAP = {
            "dept":"dept","department":"dept","branch":"dept","code":"dept",
            "count":"count","students":"count","no_of_students":"count","num":"count","strength":"count",
            "semester":"semester","sem":"semester",
            "subject":"subject","sub":"subject","paper":"subject",
        }

        rows, errors = [], []
        seen_depts   = set()
        for i, raw in enumerate(raw_rows, 1):
            norm  = {COL_MAP.get(k.strip().lower(), k.strip().lower()): v
                     for k, v in raw.items() if k.strip()}
            dept  = norm.get("dept","").strip().upper()
            count = norm.get("count","0").strip()
            sem   = norm.get("semester","3").strip()
            subj  = norm.get("subject","").strip()

            if not dept:
                errors.append(f"Row {i}: Missing department — skipped."); continue
            if dept not in DEPT_PREFIXES:
                errors.append(f"Row {i}: Unknown department '{dept}' — skipped."); continue
            if dept in seen_depts:
                errors.append(f"Row {i}: Duplicate department '{dept}' — skipped."); continue
            seen_depts.add(dept)

            try: cnt = int(float(count))
            except: errors.append(f"Row {i}: Invalid count '{count}' — skipped."); continue
            if cnt < 1:
                errors.append(f"Row {i}: Count must be ≥ 1 — skipped."); continue

            try: sem_int = int(float(sem))
            except: sem_int = 3

            rows.append({"dept":dept,"count":cnt,"semester":sem_int,"subject":subj})

        return jsonify({"rows": rows, "errors": errors})

    except Exception as ex:
        traceback.print_exc()
        return jsonify({"error": str(ex)}), 500


# ── download sample templates ──────────────────────────────────────────────────
@app.route("/api/template/<ttype>")
def download_template(ttype):
    """Serve downloadable CSV templates for both upload modes."""
    import io, csv as _csv
    buf = io.StringIO()
    w   = _csv.writer(buf)

    if ttype == "students":
        w.writerow(["register_no","name","dept","subject","semester"])
        samples = [
            ("24CSE01","Arjun Kumar","CSE","Data Structures & Algorithms","3"),
            ("23CSE32","Priya Lakshmi","CSE","Data Structures & Algorithms","3"),
            ("25ECE41","Karthik Venkat","ECE","Signals & Systems","3"),
            ("22ECE52","Divya Priya","ECE","Signals & Systems","3"),
            ("23MCE23","Surya Prakash","MECH","Thermodynamics","3"),
            ("24CVE35","Nandhini Gopal","CIVIL","Structural Analysis","3"),
            ("25EEE18","Balaji Murugan","EEE","Electrical Machines","3"),
        ]
        for s in samples: w.writerow(s)
        fname = "students_template.csv"

    elif ttype == "departments":
        w.writerow(["dept","count","semester","subject"])
        samples = [
            ("CSE","20","3","Data Structures & Algorithms"),
            ("ECE","18","3","Signals & Systems"),
            ("MECH","15","3","Thermodynamics"),
            ("CIVIL","12","3","Structural Analysis"),
            ("EEE","10","3","Electrical Machines"),
            ("AUTO","11","3","Electrical Machines"),
        ]
        for s in samples: w.writerow(s)
        fname = "departments_template.csv"
    else:
        return jsonify({"error":"Unknown template type"}), 400

    buf.seek(0)
    return send_file(io.BytesIO(buf.getvalue().encode()), as_attachment=True,
                     download_name=fname, mimetype="text/csv")


# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    has_key = bool(os.environ.get("ANTHROPIC_API_KEY","").strip())
    print("\n" + "═"*56)
    print("  Automated University Theory Examination")
    print("  Seating Arrangement System — PS09 · GCEE")
    print("═"*56)
    print(f"  URL  →  http://localhost:{port}")
    print(f"  AI   →  {'Claude Sonnet ✓' if has_key else 'Fallback parser (works without key)'}")
    print("  Stop →  Ctrl + C")
    print("═"*56+"\n")
    app.run(debug=False, port=port)
